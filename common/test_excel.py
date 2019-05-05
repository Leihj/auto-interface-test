# -*- coding: utf-8 -*-
# @File    : test_excel.PY
# @Date    : 2019/4/16-11:47
# @Author  : leihuijuan
# @Emali    : huijuan_lei@163.com

from openpyxl import load_workbook
# from common.http_api import Http_Request

class Excel_Case():

    def __init__(self):
        self.case_id=None
        self.case_title=None
        self.url=None
        self.method=None
        self.data=None
        self.expected=None
        self.actual=None
        self.result=None
        self.sql=None


class Do_Excel():

    def __init__(self,file_name,form_index):

        # try:
        #     if file_name:
                self.file_name=file_name    #文件名
                self.file_path=load_workbook(file_name)     #文件的路径
                self.sheet=self.file_path.worksheets[form_index]  #指定工作簿，用索引
                self.form_index=form_index
        # except AssertionError as e:
        #     print("{}文件不存在".format(file_name))
        # raise e


    #读取单元格的数据
    def read_cell(self,row,column):
        return self.sheet.cell(row,column).value


    #读取整个excel的数据，嵌套列表
    def read_all(self):
        list_max=[]
        for row in range(2,self.sheet.max_row+1):
            list_min = []
            for column in range(1,self.sheet.max_column+1):
                if self.sheet.cell(row,column).value:
                    list_min.append(self.sheet.cell(row,column).value)
            list_max.append(list_min)
        return list_max


    #读取指定列的数据，字典嵌套列表
    def read_column(self):
        list_out=[]
        for row in range(2,self.sheet.max_row+1):
            user={
                "case_id":self.sheet.cell(row,1).value,
                "url":self.sheet.cell(row,2).value,
                "case_data":self.sheet.cell(row,4).value,
                "expected":self.sheet.cell(row,5).value
            }
            list_out.append(user)
        return list_out


    #根据创建对象读取数据
    def create_obj_read_data(self):
        list_data=[]
        for row in range(2,self.sheet.max_row+1):
            case=Excel_Case()   #实例化Excel_Case()类
            case.case_id=self.sheet.cell(row=row,column=1).value
            case.case_title=self.sheet.cell(row=row,column=2).value
            case.url=self.sheet.cell(row=row,column=3).value
            case.method=self.sheet.cell(row=row,column=4).value
            case.data=self.sheet.cell(row=row,column=5).value
            case.expected=self.sheet.cell(row=row,column=6).value
            # case.actual=self.sheet.cell(row=row,column=7).value
            # case.result=self.sheet.cell(row=row,column=8).value
            case.sql=self.sheet.cell(row=row,column=9).value
            list_data.append(case)
        return list_data


    #填写一个单元格的数据，（回填结果）
    def write_cell(self,row,column,result):
        return self.sheet.cell(row,column,result)

    #回填多个指定单元格
    def write_result(self,row,actual,result):
        sheet=self.file_path.worksheets[self.form_index]
        sheet.cell(row=row,column=7).value=actual
        sheet.cell(row=row,column=8).value=result
        self.file_path.save(self.file_name)
        self.file_path.close()
        # self.save_excel()

    # #保存文件，如果命名新文件名则建一个新的，否则保存到原来的文件路径
    # def save_excel(self,file=None):
    #     if file:
    #         name=self.file_path.save(file)
    #
    #     else:
    #         name=self.file_path.save(self.file_name)
    #     return name
    #
    #
    # #关闭文件，每次用完excel要关闭文件
    # def close_excel(self):
    #     return self.file_path.worksheets


if __name__ == '__main__':

    from common import contants
    excel = Do_Excel(contants.case_data, 4)
    res=excel.create_obj_read_data()
    # http_res=Http_Request().http_request()
    for i in res:
        # print(i.case_id)
        # print(i.case_title)
        # print(i.url)
        print(i.__dict__)

    # excel.save_excel()